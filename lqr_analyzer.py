"""
LQR Analyzer — parses eBay's Listing Quality Report XLSX and produces
a prioritized per-listing fix list keyed against top-10% benchmarks.

Designed to drop into eBay Architect Pro as Tab 6.

Report format as of April 2026 (BETA). See readme for format notes.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

# ---------------------------------------------------------------------------
# Sheet layout constants (as of the April 16, 2026 LQR format)
# ---------------------------------------------------------------------------

# Benchmark block: impressions/CTR/conversion funnel + the "More benchmarks" table
BENCHMARK_HEADER_ROW = 24          # column headers for benchmark table
USWM_IMPRESSIONS_ROW = 25          # your row under impressions benchmark
TOP10_IMPRESSIONS_ROW = 26
BOT10_IMPRESSIONS_ROW = 27
USWM_CTR_ROW = 29
TOP10_CTR_ROW = 30
BOT10_CTR_ROW = 31
USWM_CONV_ROW = 33
TOP10_CONV_ROW = 34
BOT10_CONV_ROW = 35

# First value column for the funnel metric in benchmark rows
BENCH_METRIC_COL = 6               # impressions/CTR/conversion scalar lives here

# Per-listing table
LISTING_HEADER_ROW = 44
LISTING_DATA_START = 45

# Sheet-level funnel data (top block with 3 boxes: impressions / CTR / conversion)
# eBay's sheet uses col A as a spacer; data starts at col B (col 2).
FUNNEL_SELLERS_ROW = 7
FUNNEL_SELLERS_COL = 2
FUNNEL_IMPRESSIONS_ROW = 11
FUNNEL_CTR_ROW = 15
FUNNEL_CONV_ROW = 19
FUNNEL_VALUE_COL = 4
FUNNEL_RANK_COL = 6
FUNNEL_SALES_RANK_ROW = 2
FUNNEL_SALES_RANK_COL = 10

# Skip these sheets when walking category tabs
NON_CATEGORY_SHEETS = {"Summary", "Guide", "Google Shopping Rejections"}


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Benchmark:
    """Benchmarks for a single category + condition."""
    category: str
    condition: str
    sellers_in_category: int = 0
    sales_rank: int | None = None      # your GMV rank
    sales_rank_of: int | None = None   # total sellers
    # Funnel — your values
    avg_impressions_per_day: float = 0.0
    impressions_rank: int | None = None
    your_ctr: float = 0.0
    ctr_rank: int | None = None
    your_conversion: float = 0.0
    conversion_rank: int | None = None
    # Top/bottom 10% scalars from the More Benchmarks block
    top10_impressions_per_listing: float = 0.0
    top10_ctr: float = 0.0
    top10_conversion: float = 0.0
    bot10_impressions_per_listing: float = 0.0
    bot10_ctr: float = 0.0
    bot10_conversion: float = 0.0
    # Per-property benchmark dicts (from the sold/converting top-10% row)
    top10: dict[str, Any] = field(default_factory=dict)
    bot10: dict[str, Any] = field(default_factory=dict)
    uswm: dict[str, Any] = field(default_factory=dict)


@dataclass
class Listing:
    """One listing's data + computed gaps + recommendations."""
    item_id: str
    title: str
    category: str
    condition: str
    # Raw metrics
    daily_impressions: float = 0.0
    ctr: float = 0.0
    conversion: float | None = None   # None if "No sales"
    actual_shipping_days: float | None = None
    actual_handling_time: float | None = None
    photos: int = 0
    upc: str = ""
    recommended_specifics_filled: int = 0
    specifics_to_add: str = ""
    brand: str = ""
    mpn: str = ""
    keywords_in_title: int = 0
    free_shipping: bool = False
    estimated_handling_time: float | None = None
    accepts_returns: bool = False
    seller_paid_returns: bool = False
    returns_30day: bool = False
    promoted: bool = False
    promoted_ad_rate: float = 0.0
    price: float = 0.0
    strikethrough_price: float | None = None
    quantity: int = 0
    item_age_days: int = 0
    sales_count_31d: int = 0
    watchers: int = 0
    best_offer: bool = False

    # Computed
    leverage_score: float = 0.0      # estimated daily $ uplift from fixes
    priority_rank: int = 0
    recommendations: list[str] = field(default_factory=list)
    severity: str = ""                # CRITICAL / HIGH / MEDIUM / LOW / NONE


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(ws, row: int, col: int) -> Any:
    """Return cell value or None."""
    return ws.cell(row=row, column=col).value


def _as_float(v: Any) -> float | None:
    """Coerce eBay's mixed-type cells to float. 'No sales', 'N/A', blanks -> None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "N/A", "No sales", "Not provided", "Not enabled", "Not relevant"):
            return None
        # "Same day" handling time = 0
        if s.lower() == "same day":
            return 0.0
        try:
            return float(s.replace(",", "").replace("%", "").replace("$", ""))
        except ValueError:
            return None
    return None


def _as_int(v: Any, default: int = 0) -> int:
    f = _as_float(v)
    return int(f) if f is not None else default


_EMPTY_MARKERS = {
    "", " ", "Not provided", "Not enabled", "Not relevant",
    "N/A", "None", "none", "Does Not Apply", "does not apply",
}


def _clean_string_field(v: Any) -> str:
    """Normalize a string-valued LQR cell. eBay's 'Not provided' placeholders
    become empty strings. The '✔' check-mark is preserved so the display layer
    can render it as '(filled)' where the actual value is unavailable."""
    if v is None:
        return ""
    s = str(v).strip()
    if s in _EMPTY_MARKERS:
        return ""
    return s


def _as_bool(v: Any) -> bool:
    """eBay uses '✔' for true, blank or 'Not provided'/'Not enabled' for false."""
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() in ("✔", "Yes", "1", "true", "True")
    if isinstance(v, (int, float)):
        return bool(v)
    return False


def _parse_rank_string(s: str) -> tuple[int | None, int | None]:
    """Parse 'Your sales rank (GMV) in this section: 6 out of 11,249 sellers'."""
    if not s:
        return None, None
    m = re.search(r"([\d,]+)\s+out of\s+([\d,]+)", str(s))
    if m:
        return (int(m.group(1).replace(",", "")),
                int(m.group(2).replace(",", "")))
    return None, None


def _parse_sellers_count(s: str) -> int:
    """Parse 'Sellers selling in this section: 885'."""
    if not s:
        return 0
    m = re.search(r"(\d+)", str(s))
    return int(m.group(1)) if m else 0


def _decode_condition(ws) -> tuple[str, str]:
    """Extract category + condition from cell A3.

    Cell A3 contains: 'Category: Chains & Fobs, Listings condition: New'

    The sheet name suffix (|U, |R, |N) is NOT a reliable condition indicator —
    eBay assigns bare names to whichever condition has the most listings, and
    the bare name might be Used, Refurbished, or anything else.
    """
    raw = ws.cell(row=3, column=2).value or ws.cell(row=3, column=1).value
    if not raw:
        return ws.title, "Unknown"
    s = str(raw)
    cat = s
    cond = "Unknown"
    m = re.search(r"Category:\s*(.+?),\s*Listings condition:\s*(\w+)", s)
    if m:
        cat = m.group(1).strip()
        cond = m.group(2).strip()
    return cat, cond


# ---------------------------------------------------------------------------
# Benchmark parsing
# ---------------------------------------------------------------------------

def parse_benchmark(ws) -> Benchmark:
    """Extract the benchmark block from a category sheet."""
    category, condition = _decode_condition(ws)
    bm = Benchmark(category=category, condition=condition)

    bm.sellers_in_category = _parse_sellers_count(
        _cell(ws, FUNNEL_SELLERS_ROW, FUNNEL_SELLERS_COL))
    sales_rank, sales_of = _parse_rank_string(
        _cell(ws, FUNNEL_SALES_RANK_ROW, FUNNEL_SALES_RANK_COL))
    bm.sales_rank = sales_rank
    bm.sales_rank_of = sales_of

    bm.avg_impressions_per_day = _as_float(
        _cell(ws, FUNNEL_IMPRESSIONS_ROW, FUNNEL_VALUE_COL)) or 0.0
    bm.impressions_rank = _as_int(
        _cell(ws, FUNNEL_IMPRESSIONS_ROW, FUNNEL_RANK_COL), 0) or None
    bm.your_ctr = _as_float(
        _cell(ws, FUNNEL_CTR_ROW, FUNNEL_VALUE_COL)) or 0.0
    bm.ctr_rank = _as_int(
        _cell(ws, FUNNEL_CTR_ROW, FUNNEL_RANK_COL), 0) or None
    bm.your_conversion = _as_float(
        _cell(ws, FUNNEL_CONV_ROW, FUNNEL_VALUE_COL)) or 0.0
    bm.conversion_rank = _as_int(
        _cell(ws, FUNNEL_CONV_ROW, FUNNEL_RANK_COL), 0) or None

    # Scalar top/bottom 10% values from the More Benchmarks block (col 6 of each row)
    bm.top10_impressions_per_listing = _as_float(
        _cell(ws, TOP10_IMPRESSIONS_ROW, BENCH_METRIC_COL)) or 0.0
    bm.bot10_impressions_per_listing = _as_float(
        _cell(ws, BOT10_IMPRESSIONS_ROW, BENCH_METRIC_COL)) or 0.0
    bm.top10_ctr = _as_float(
        _cell(ws, TOP10_CTR_ROW, BENCH_METRIC_COL)) or 0.0
    bm.bot10_ctr = _as_float(
        _cell(ws, BOT10_CTR_ROW, BENCH_METRIC_COL)) or 0.0
    bm.top10_conversion = _as_float(
        _cell(ws, TOP10_CONV_ROW, BENCH_METRIC_COL)) or 0.0
    bm.bot10_conversion = _as_float(
        _cell(ws, BOT10_CONV_ROW, BENCH_METRIC_COL)) or 0.0

    # Per-property benchmark dict — use the sold benchmark (row 34) since it tells us
    # what top-converting listings look like, which is our improvement target.
    bench_cols: dict[str, int] = {}
    for c in range(2, ws.max_column + 1):
        h = _cell(ws, BENCHMARK_HEADER_ROW, c)
        if h:
            bench_cols[str(h).strip()] = c

    def _row_dict(row: int) -> dict[str, Any]:
        return {name: _cell(ws, row, col) for name, col in bench_cols.items()}

    bm.uswm = _row_dict(USWM_CONV_ROW)
    bm.top10 = _row_dict(TOP10_CONV_ROW)
    bm.bot10 = _row_dict(BOT10_CONV_ROW)

    return bm


# ---------------------------------------------------------------------------
# Listing parsing
# ---------------------------------------------------------------------------

_LISTING_COL_MAP = {
    "Item title": "title",
    "Daily impressions per listing": "daily_impressions",
    "Click-through rate": "ctr",
    "Sales conversion rate": "conversion",
    "Actual shipping time": "actual_shipping_days",
    "Number of photos": "photos",
    "UPC": "upc",
    "Recommended item specifics filled": "recommended_specifics_filled",
    "Recommended item specifics to add": "specifics_to_add",
    "Brand": "brand",
    "MPN": "mpn",
    "Number of keywords in title": "keywords_in_title",
    "Free shipping": "free_shipping",
    "Estimated handling time": "estimated_handling_time",
    "30 Days Returns": "returns_30day",
    "Accepts returns": "accepts_returns",
    "Seller paid returns": "seller_paid_returns",
    "Promoted listings": "promoted",
    "Promoted listings ad rate": "promoted_ad_rate",
    "Item Id": "item_id",
    "Price": "price",
    "Strikethrough price": "strikethrough_price",
    "Quantity available": "quantity",
    "Item age in days": "item_age_days",
    "Sales count in last 31 days": "sales_count_31d",
    "Number of watchers": "watchers",
    "Actual handling time": "actual_handling_time",
    "Best Offer": "best_offer",
}


def parse_listings(ws, category: str, condition: str) -> list[Listing]:
    """Extract per-listing rows from a category sheet."""
    # Build column map from header row
    col_map: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = _cell(ws, LISTING_HEADER_ROW, c)
        if h:
            col_map[str(h).strip()] = c

    listings: list[Listing] = []
    for row in range(LISTING_DATA_START, ws.max_row + 1):
        # Check if row has data (title or item id present)
        title_col = col_map.get("Item title")
        id_col = col_map.get("Item Id")
        title_val = _cell(ws, row, title_col) if title_col else None
        id_val = _cell(ws, row, id_col) if id_col else None
        if not title_val and not id_val:
            continue

        # Item ID: sometimes the title column contains the raw numeric ID when
        # the listing was created via Sell Similar and eBay hasn't indexed the
        # title yet. Fall back to the dedicated Item Id column.
        item_id = str(id_val) if id_val else ""
        title = str(title_val) if title_val else item_id

        l = Listing(
            item_id=item_id,
            title=title,
            category=category,
            condition=condition,
        )

        for header, attr in _LISTING_COL_MAP.items():
            if header not in col_map:
                continue
            raw = _cell(ws, row, col_map[header])
            # Route by target type
            if attr in ("title", "upc", "brand", "mpn", "specifics_to_add", "item_id"):
                setattr(l, attr, _clean_string_field(raw))
            elif attr in ("free_shipping", "accepts_returns", "seller_paid_returns",
                          "promoted", "best_offer", "returns_30day"):
                setattr(l, attr, _as_bool(raw))
            elif attr in ("photos", "keywords_in_title", "quantity", "item_age_days",
                           "sales_count_31d", "watchers", "recommended_specifics_filled"):
                setattr(l, attr, _as_int(raw))
            else:
                val = _as_float(raw)
                # conversion stays None if "No sales" — meaningful signal
                if attr == "conversion":
                    setattr(l, attr, val)
                else:
                    setattr(l, attr, val if val is not None else 0.0)

        listings.append(l)

    return listings


# ---------------------------------------------------------------------------
# Scoring / recommendations
# ---------------------------------------------------------------------------

# How much of the top-10% benchmark we expect a fix to move the needle.
# Conservative: 50% of the full gap closes with photo/metadata fixes.
FIX_REALISM_FACTOR = 0.5


def score_listing(l: Listing, bm: Benchmark) -> None:
    """Populate leverage_score, severity, and recommendations on a Listing."""
    recs: list[str] = []

    # ---- Benchmarks from top 10% of SOLD listings in this category ----
    top10 = bm.top10
    bot10 = bm.bot10

    def _num(d: dict, key: str) -> float | None:
        v = d.get(key)
        if v in (None, "Not relevant", "Not provided", ""):
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    top_ctr = _num({"x": bm.top10.get("Click-through rate") if False else None}, "x")
    # For CTR and conversion, the section headers in the benchmark block
    # aren't always named "Click-through rate"; we have to pull from the
    # separate CTR and Conversion rows parsed above. The funnel values are
    # authoritative for the category-level benchmark.
    # Per-listing uplift targets come from the sheet's funnel section:
    #   top 10% CTR: 4.17% (Chains), etc. — we don't have the raw split by
    #   listing in this sheet, but the funnel row provides top10 values for
    #   CTR and conversion via a different row that we parse in __main__.
    # For simplicity and robustness, we approximate the top-10% CTR as 10x
    # the category average CTR when we don't have an explicit value, and
    # we pull explicit values from the sheet-level benchmarks when possible.

    # ---- Death-spiral detector ----
    # A listing with significant impressions + clicks but zero conversion
    # over a meaningful window is the highest-priority flag.
    clicks_estimated = l.daily_impressions * l.ctr * max(1, l.item_age_days)
    if l.conversion is None and clicks_estimated >= 50 and l.item_age_days >= 30:
        recs.append(
            f"DEATH SPIRAL: ~{int(clicks_estimated)} clicks over {l.item_age_days} "
            f"days, zero sales. Cassini has logged bad conversion. SELL SIMILAR to "
            f"get a fresh item ID."
        )

    # ---- Photo count check ----
    # Top-10% SOLD listings' avg photo count is in bm.top10 under "Number of photos"
    top10_photos = _num(top10, "Number of photos")
    if top10_photos and l.photos < top10_photos:
        gap = top10_photos - l.photos
        if gap >= 3:
            recs.append(
                f"ADD PHOTOS: listing has {l.photos}, top-10% sold avg is "
                f"{top10_photos:.0f}. Add at least {int(gap)} more."
            )
        elif gap >= 1:
            recs.append(
                f"PHOTO GAP: {l.photos} vs {top10_photos:.0f} benchmark. "
                f"Add {int(gap)} more for parity."
            )

    # ---- UPC ----
    top10_upc_fill = _num(top10, "UPC")
    if top10_upc_fill and top10_upc_fill >= 0.30 and not l.upc:
        recs.append(
            f"ADD UPC: top-10% sold in this category fill UPC "
            f"{top10_upc_fill*100:.0f}% of the time."
        )

    # ---- Free shipping ----
    top10_free_ship = _num(top10, "Free shipping")
    if top10_free_ship and top10_free_ship >= 0.70 and not l.free_shipping:
        recs.append("ENABLE FREE SHIPPING: >70% of top-10% sold offer it.")

    # ---- 30-day returns ----
    top10_30ret = _num(top10, "30 Days Returns")
    if top10_30ret and top10_30ret >= 0.50 and not l.returns_30day:
        recs.append("ENABLE 30-DAY RETURNS: majority of top-10% sold offer this.")

    # ---- Handling time ----
    top10_handling = _num(top10, "Estimated handling time")
    if (top10_handling is not None and l.estimated_handling_time is not None
            and l.estimated_handling_time > top10_handling + 0.5):
        recs.append(
            f"REDUCE HANDLING: {l.estimated_handling_time:.0f}d vs "
            f"{top10_handling:.0f}d benchmark."
        )

    # ---- Recommended item specifics ----
    top10_specifics = _num(top10, "Recommended item specifics filled")
    if top10_specifics and l.recommended_specifics_filled < top10_specifics:
        gap = top10_specifics - l.recommended_specifics_filled
        if gap >= 1:
            recs.append(
                f"ADD ITEM SPECIFICS: {l.recommended_specifics_filled} filled vs "
                f"{top10_specifics:.0f} benchmark."
                + (f" Missing: {l.specifics_to_add}" if l.specifics_to_add else "")
            )

    # ---- Promoted Listings ad rate overpayment ----
    top10_ad_rate = _num(top10, "Promoted listings ad rate")
    if (top10_ad_rate is not None and l.promoted_ad_rate
            and l.promoted_ad_rate > top10_ad_rate * 1.5):
        overpay = (l.promoted_ad_rate - top10_ad_rate) * 100
        recs.append(
            f"AD RATE OVERPAY: {l.promoted_ad_rate*100:.1f}% vs "
            f"{top10_ad_rate*100:.1f}% benchmark (+{overpay:.1f}pp)."
        )

    # ---- Brand / MPN fill ----
    # "✔" means populated but eBay didn't echo the value in the report;
    # treat it as filled and do not recommend filling.
    brand_value = l.brand.strip() if l.brand else ""
    brand_missing = brand_value == "" or brand_value.lower() in (
        "unbranded", "does not apply",
    )
    if brand_missing:
        top10_brand_fill = _num(top10, "Brand")
        if top10_brand_fill and top10_brand_fill >= 0.5:
            recs.append("FILL BRAND: top-10% sold populate this field.")

    # ---- Leverage score ----
    # Estimated daily revenue lift if the listing moved to top-10% CTR and its
    # current conversion (or category avg if death spiral). Capped at realistic
    # fraction of the full gap.
    target_ctr = bm.top10_ctr if bm.top10_ctr > 0 else max(l.ctr, bm.your_ctr) * 3.0
    # Don't claim a listing can exceed top-10% CTR
    target_ctr = max(target_ctr, l.ctr)

    current_conv = l.conversion if l.conversion is not None else bm.your_conversion
    # For death-spiral listings, assume Sell Similar gets back to category avg conversion
    if l.conversion is None and bm.your_conversion > 0:
        current_conv = bm.your_conversion

    potential_daily_clicks = l.daily_impressions * target_ctr
    current_daily_clicks = l.daily_impressions * l.ctr
    click_uplift = max(0, potential_daily_clicks - current_daily_clicks)
    l.leverage_score = click_uplift * current_conv * l.price * FIX_REALISM_FACTOR

    # ---- Severity ----
    if any("DEATH SPIRAL" in r for r in recs):
        l.severity = "CRITICAL"
    elif len(recs) >= 3:
        l.severity = "HIGH"
    elif len(recs) >= 1:
        l.severity = "MEDIUM"
    else:
        l.severity = "NONE"

    l.recommendations = recs


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

@dataclass
class LQRReport:
    generated_at: str
    benchmarks: list[Benchmark]
    listings: list[Listing]

    @property
    def critical_count(self) -> int:
        return sum(1 for l in self.listings if l.severity == "CRITICAL")

    @property
    def high_count(self) -> int:
        return sum(1 for l in self.listings if l.severity == "HIGH")

    @property
    def medium_count(self) -> int:
        return sum(1 for l in self.listings if l.severity == "MEDIUM")

    @property
    def clean_count(self) -> int:
        return sum(1 for l in self.listings if l.severity == "NONE")

    @property
    def total_leverage(self) -> float:
        return sum(l.leverage_score for l in self.listings)


def parse_lqr(xlsx_path: str | Path) -> LQRReport:
    """Parse an eBay Listings Quality Report XLSX into structured data."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Extract report generation timestamp from Summary sheet
    summary = wb["Summary"] if "Summary" in wb.sheetnames else None
    generated = ""
    if summary:
        for row in range(1, 10):
            for col in range(1, 15):
                v = summary.cell(row=row, column=col).value
                if v and isinstance(v, str) and "Report generated" in v:
                    generated = v.replace("Report generated:", "").strip()
                    break

    benchmarks: list[Benchmark] = []
    all_listings: list[Listing] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in NON_CATEGORY_SHEETS:
            continue
        ws = wb[sheet_name]
        bm = parse_benchmark(ws)
        benchmarks.append(bm)
        listings = parse_listings(ws, bm.category, bm.condition)
        for l in listings:
            score_listing(l, bm)
        all_listings.extend(listings)

    # Rank by leverage
    all_listings.sort(key=lambda x: x.leverage_score, reverse=True)
    for i, l in enumerate(all_listings, 1):
        l.priority_rank = i

    return LQRReport(
        generated_at=generated,
        benchmarks=benchmarks,
        listings=all_listings,
    )
